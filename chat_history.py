import json
import os
import sys
from datetime import datetime as dt
from pathlib import Path
from uuid import uuid4

import certifi
import pytz
import questionary
from bson.objectid import ObjectId
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook

from tools.azure_env import AzureEnv
from tools.misc import format_datetime

timezone = pytz.timezone("UTC")


class ChatHistory:
    def __init__(self, uri, db_name, chat_history_collection, user_collection, feedback_collection):
        self.client = AsyncIOMotorClient(uri, tlsCAFile=certifi.where())
        self.db = self.client[db_name]
        self.chat_history_collection = self.db[chat_history_collection]
        self.user_collection = self.db[user_collection]
        self.feedback_collection = self.db[feedback_collection]

    async def remove_thumb_properties(self):
        self.chat_history_collection.update_many({}, {"$unset": {"thumbs_up": 1, "thumbs_down": 1}})

    async def transfer_thumbs_to_feedback(self):
        chat_history = self.chat_history_collection.find({"$or": [{"thumbs_up": {"$eq": 1}}, {"thumbs_down": {"$eq": 1}}]})
        async for ch in chat_history:
            print(ch["_id"], ch["thumbs_up"], ch["thumbs_down"])

            await self.feedback_collection.insert_one(
                {
                    "shard_id": str(uuid4()),
                    "chat_id": ObjectId(ch["_id"]),
                    "thumbs_up": ch["thumbs_up"],
                    "thumbs_down": ch["thumbs_down"],
                    "feedback_choices": [],
                    "feedback_details": "",
                    "created_at": ch["created_at"],
                    "timezone": ch["timezone"],
                }
            )

    async def update_chat_history_with_feedback_id(self):
        feedbacks = self.feedback_collection.find({})
        async for feedback in feedbacks:
            self.chat_history_collection.update_many({"_id": feedback["chat_id"]}, {"$set": {"feedback_id": feedback["_id"]}})

    async def update_feedback(self):
        chat_history = self.chat_history_collection.find({"feedback_id": {"$exists": True}})
        async for ch in chat_history:
            self.feedback_collection.update_many({"_id": ch["feedback_id"]}, {"$set": {"created_at": ch["created_at"], "timezone": ch["timezone"]}})

    async def find_questions(self):
        chat_history = await self.chat_history_collection.find({"question": {"$regex": ".*disregard all previous instructions.*"}}).to_list(None)
        print(chat_history[0]["created_at"])

    async def get_users(self, start_date: str, end_date: str, sort_by: str, order_by: str, file_type: str = "json"):
        start_date_dt, end_date_dt = format_datetime(start_date, end_date)

        sort_direction = 1 if order_by == "ASC" else -1
        paginated_stage = [{"$sort": {sort_by.lower(): sort_direction}}]

        users = await self.chat_history_collection.aggregate(
            [
                {
                    "$addFields": {
                        "created_at_localized": {
                            "$dateToString": {
                                "date": "$created_at",
                                "format": "%Y-%m-%d %H:%M:%S",
                                "timezone": {
                                    "$switch": {
                                        "branches": [
                                            {"case": {"$eq": ["$timezone", "Europe/Kyiv"]}, "then": "Europe/Kiev"},
                                            {"case": {"$eq": ["$timezone", "America/Ciudad_Juarez"]}, "then": "America/Denver"},
                                        ],
                                        "default": "$timezone",
                                    }
                                },
                            }
                        }
                    }
                },
                {"$match": {"created_at_localized": {"$gte": start_date_dt, "$lte": end_date_dt}}},
                {"$lookup": {"from": "users", "localField": "user_id", "foreignField": "user_id", "as": "user"}},
                {"$unwind": "$user"},
                {
                    "$addFields": {
                        "last_active_at_localized": {
                            "$dateToString": {
                                "date": "$user.last_active_at",
                                "format": "%Y-%m-%d %H:%M:%S",
                                "timezone": {
                                    "$switch": {
                                        "branches": [
                                            {"case": {"$eq": ["$timezone", "Europe/Kyiv"]}, "then": "Europe/Kiev"},
                                            {"case": {"$eq": ["$timezone", "America/Ciudad_Juarez"]}, "then": "America/Denver"},
                                        ],
                                        "default": "$timezone",
                                    }
                                },
                            }
                        }
                    }
                },
                {"$lookup": {"from": "feedback", "localField": "feedback_id", "foreignField": "_id", "as": "feedback"}},
                {"$unwind": {"path": "$feedback", "preserveNullAndEmptyArrays": True}},
                {
                    "$group": {
                        "_id": "$user.user_id",
                        "name": {"$first": "$user.name"},
                        "total_questions": {"$sum": 1},
                        "total_answers_no_citations": {
                            "$sum": {"$cond": {"if": {"$or": [{"$eq": ["$citations", []]}, {"$eq": ["$citations", ""]}]}, "then": 1, "else": 0}}
                        },
                        "total_thumbs_up": {"$sum": "$feedback.thumbs_up"},
                        "total_thumbs_down": {"$sum": "$feedback.thumbs_down"},
                        "DateSet": {"$addToSet": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}}},
                        "created_at": {"$first": "$created_at_localized"},
                        "last_active_at": {"$first": "$last_active_at_localized"},
                        "timezone": {"$first": "$timezone"},
                    }
                },
                {
                    "$project": {
                        "_id": 0,
                        "user_id": "$_id",
                        "name": 1,
                        "total_questions": 1,
                        "total_answers_no_citations": 1,
                        "total_thumbs_up": 1,
                        "total_thumbs_down": 1,
                        "total_visits": {"$cond": {"if": {"$gte": [{"$size": "$DateSet"}, 1]}, "then": {"$size": "$DateSet"}, "else": 0}},
                        # "created_at": 1,
                        # "last_active_at": 1,
                        "timezone": 1,
                    }
                },
                {"$facet": {"paginatedResults": paginated_stage, "totalCount": [{"$count": "user_id"}]}},
            ]
        ).to_list(length=None)

        if len(users[0]["totalCount"]) == 0:
            return {"items": [], "total_count": 0}

        total_count = users[0]["totalCount"][0]["user_id"]
        users = users[0]["paginatedResults"]

        # for user in users:
        #     for key, value in user.items():
        #         if key in ["thoughts", "answer"]:
        #             continue

        #         print(key, ":", value)
        #     print("")

        if file_type == "json":
            with open("users.json", "w+", encoding="utf-8") as f:
                json.dump(users, f)
        elif file_type == "xlsx":
            workbook = Workbook()

            sheet = workbook.active
            sheet.title = "Users"

            # Write headers (keys from the first dictionary)
            headers = users[0].keys()
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_index, value=header)

            # Write rows (values from each dictionary)
            for row_index, record in enumerate(users, start=2):  # Start writing from row 2
                for col_index, key in enumerate(headers, start=1):
                    sheet.cell(row=row_index, column=col_index, value=record[key])

            output_file = "Users Q4.xlsx"
            workbook.save(output_file)


if __name__ == "__main__":
    env = questionary.select("Which environment?", choices=["prod", "dev"]).ask()
    brand = questionary.select("Which brand?", choices=["clo3d", "closet", "md"]).ask()
    # task = questionary.select("What task?", choices=["Get Posts", "Upload"]).ask()

    azure_env = AzureEnv(env, brand)

    ch = ChatHistory(azure_env.URI, azure_env.DB_NAME, azure_env.COLLECTION_NAME, azure_env.COLLECTION_USERS, azure_env.COLLECTION_FEEDBACK)
    loop = ch.client.get_io_loop()

    start_date = "2024-10-01 00:00:00"
    end_date = "2024-12-31 23:59:59"
    # end_date = dt.now().strftime("%Y-%m-%d %H:%M:%S")

    loop.run_until_complete(ch.get_users(start_date, end_date, "user_id", "asc", "xlsx"))
